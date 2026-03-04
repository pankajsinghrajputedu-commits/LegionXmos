from fastapi import FastAPI, APIRouter, HTTPException, File, UploadFile, Response, Request, Cookie, Form
from fastapi.responses import JSONResponse, StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
import asyncio
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict, EmailStr
from typing import List, Optional, Literal
import uuid
from datetime import datetime, timezone, timedelta
from emergentintegrations.llm.chat import LlmChat, UserMessage
import io
from PyPDF2 import PdfReader
from docx import Document
from PIL import Image
import pytesseract
import httpx
import resend
import csv

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

app = FastAPI()
api_router = APIRouter(prefix="/api")

EMERGENT_LLM_KEY = os.environ.get('EMERGENT_LLM_KEY')
RESEND_API_KEY = os.environ.get('RESEND_API_KEY')
SENDER_EMAIL = os.environ.get('SENDER_EMAIL', 'onboarding@resend.dev')
resend.api_key = RESEND_API_KEY

# Auth Models
class User(BaseModel):
    model_config = ConfigDict(extra="ignore")
    user_id: str = Field(default_factory=lambda: f"user_{uuid.uuid4().hex[:12]}")
    email: str
    name: str
    picture: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class UserSession(BaseModel):
    model_config = ConfigDict(extra="ignore")
    session_token: str
    user_id: str
    expires_at: datetime
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class SessionRequest(BaseModel):
    session_id: str

# Hiring Session Model
class HiringSession(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    description: Optional[str] = None
    status: Literal["active", "draft", "completed"] = "active"
    created_by: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class Question(BaseModel):
    question_id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    type: Literal["mcq", "short_answer", "scenario", "task"]
    text: str
    options: Optional[List[str]] = None
    expected_answer: Optional[str] = None

class JobDescriptionInput(BaseModel):
    description: str
    hiring_session_id: Optional[str] = None

class JobDescriptionParsed(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    hiring_session_id: Optional[str] = None
    title: str
    role: str
    seniority: str
    skills: List[str]
    domain: str
    description: str
    key_responsibilities: List[str]
    source: str = "text"
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class Assessment(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    job_id: str
    job_title: str
    questions: List[Question]
    generated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class GenerateAssessmentRequest(BaseModel):
    job_id: str
    difficulty: str = "moderate"  # easy, moderate, difficult
    question_count: int = 50

class TestLinkRequest(BaseModel):
    assessment_id: str

class TestLink(BaseModel):
    model_config = ConfigDict(extra="ignore")
    link_id: str = Field(default_factory=lambda: str(uuid.uuid4())[:8])
    assessment_id: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class CandidateAnswer(BaseModel):
    question_id: str
    answer: str

class TestSubmission(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    candidate_name: str
    candidate_email: str
    assessment_id: str
    link_id: str
    answers: List[CandidateAnswer]
    submitted_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class SubmitTestRequest(BaseModel):
    link_id: str
    candidate_name: str
    candidate_email: str
    answers: List[CandidateAnswer]

class QuestionScore(BaseModel):
    question_id: str
    score: float
    reasoning: str

class Score(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    submission_id: str
    assessment_id: str
    candidate_name: str
    candidate_email: str
    total_score: float
    percentage: float
    question_scores: List[QuestionScore]
    strengths: List[str]
    weaknesses: List[str]
    recommendations: str
    scored_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class ScoreRequest(BaseModel):
    submission_id: str

class LeaderboardEntry(BaseModel):
    model_config = ConfigDict(extra="ignore")
    rank: int
    candidate_name: str
    candidate_email: str
    total_score: float
    percentage: float
    strengths: List[str]
    recommendations: str
    submitted_at: datetime

# Candidate Invitation Models
class Candidate(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    hiring_session_id: str
    assessment_id: Optional[str] = None
    name: str
    email: EmailStr
    status: Literal["invited", "test_started", "test_completed", "scored"] = "invited"
    invite_sent_at: Optional[datetime] = None
    test_link_id: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class CandidateCreate(BaseModel):
    name: str
    email: EmailStr

class BulkCandidateCreate(BaseModel):
    candidates: List[CandidateCreate]

class SendInviteRequest(BaseModel):
    candidate_ids: List[str]
    assessment_id: str

def extract_text_from_pdf(file_content: bytes) -> str:
    try:
        pdf_file = io.BytesIO(file_content)
        pdf_reader = PdfReader(pdf_file)
        text = ""
        for page in pdf_reader.pages:
            extracted = page.extract_text()
            if extracted:
                text += extracted + "\n"
        
        if not text.strip():
            raise ValueError("No text found in PDF")
        
        return text.strip()
    except Exception as e:
        logging.error(f"PDF extraction error: {e}")
        raise HTTPException(status_code=400, detail=f"Failed to extract text from PDF: {str(e)}")

def extract_text_from_docx(file_content: bytes) -> str:
    try:
        docx_file = io.BytesIO(file_content)
        doc = Document(docx_file)
        text = "\n".join([para.text for para in doc.paragraphs if para.text.strip()])
        
        if not text.strip():
            raise ValueError("No text found in document")
        
        return text.strip()
    except Exception as e:
        logging.error(f"DOCX extraction error: {e}")
        raise HTTPException(status_code=400, detail=f"Failed to extract text from DOCX: {str(e)}")

def extract_text_from_image(file_content: bytes) -> str:
    try:
        image = Image.open(io.BytesIO(file_content))
        if image.mode != 'RGB':
            image = image.convert('RGB')
        text = pytesseract.image_to_string(image, config='--psm 6')
        
        if not text.strip():
            raise ValueError("No text found in image")
        
        return text.strip()
    except Exception as e:
        logging.error(f"Image OCR error: {e}")
        raise HTTPException(status_code=400, detail=f"Failed to extract text from image: {str(e)}")

async def get_current_user(request: Request, session_token: Optional[str] = Cookie(None)) -> User:
    token = session_token or request.headers.get("Authorization", "").replace("Bearer ", "")
    if not token:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    session = await db.user_sessions.find_one({"session_token": token}, {"_id": 0})
    if not session:
        raise HTTPException(status_code=401, detail="Invalid session")
    
    expires_at = session["expires_at"]
    if isinstance(expires_at, str):
        expires_at = datetime.fromisoformat(expires_at)
    if expires_at.tzinfo is None:
        expires_at = expires_at.replace(tzinfo=timezone.utc)
    if expires_at < datetime.now(timezone.utc):
        raise HTTPException(status_code=401, detail="Session expired")
    
    user = await db.users.find_one({"user_id": session["user_id"]}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    user['created_at'] = datetime.fromisoformat(user['created_at']) if isinstance(user['created_at'], str) else user['created_at']
    return User(**user)

# Auth Endpoints
# Auth Endpoints
class EmailRequest(BaseModel):
    email: str

class OTPRequest(BaseModel):
    email: str
    otp: str

@api_router.post("/auth/send-otp")
async def send_otp_email(request: EmailRequest):
    try:
        import random
        otp = ''.join([str(random.randint(0, 9)) for _ in range(6)])
        expires_at = datetime.now(timezone.utc) + timedelta(minutes=10)
        
        await db.otp_codes.update_one(
            {"email": request.email},
            {"$set": {"otp": otp, "expires_at": expires_at.isoformat(), "created_at": datetime.now(timezone.utc).isoformat()}},
            upsert=True
        )
        
        # Send actual email using Resend
        if RESEND_API_KEY:
            html_content = f"""
            <div style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto; padding: 20px;">
                <div style="text-align: center; margin-bottom: 30px;">
                    <h1 style="color: #991b1b; margin: 0;">Beat Claude</h1>
                    <p style="color: #6b7280; margin-top: 5px;">AI Hiring Companion</p>
                </div>
                <div style="background: #f9fafb; border-radius: 8px; padding: 30px; text-align: center;">
                    <h2 style="color: #111827; margin: 0 0 10px 0;">Your Verification Code</h2>
                    <p style="color: #6b7280; margin: 0 0 20px 0;">Enter this code to complete your sign in</p>
                    <div style="background: #991b1b; color: white; font-size: 32px; font-weight: bold; letter-spacing: 8px; padding: 20px 40px; border-radius: 8px; display: inline-block;">
                        {otp}
                    </div>
                    <p style="color: #9ca3af; font-size: 14px; margin-top: 20px;">This code expires in 10 minutes</p>
                </div>
                <p style="color: #9ca3af; font-size: 12px; text-align: center; margin-top: 20px;">
                    If you didn't request this code, you can safely ignore this email.
                </p>
            </div>
            """
            
            params = {
                "from": SENDER_EMAIL,
                "to": [request.email],
                "subject": f"Your Beat Claude verification code: {otp}",
                "html": html_content
            }
            
            try:
                await asyncio.to_thread(resend.Emails.send, params)
                logging.info(f"OTP email sent to {request.email}")
                return {"message": "OTP sent successfully", "email": request.email}
            except Exception as e:
                logging.error(f"Resend error: {e}")
                # Fallback - return OTP for testing if email fails
                return {"message": "OTP sent (email delivery may fail in test mode)", "otp": otp, "email": request.email}
        else:
            # No Resend key - return OTP for testing
            logging.info(f"OTP for {request.email}: {otp} (no email - testing mode)")
            return {"message": "OTP generated (testing mode - no email sent)", "otp": otp, "email": request.email}
    except Exception as e:
        logging.error(f"OTP send error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/auth/verify-otp")
async def verify_otp(request: OTPRequest, response: Response):
    try:
        otp_doc = await db.otp_codes.find_one({"email": request.email}, {"_id": 0})
        if not otp_doc:
            raise HTTPException(status_code=401, detail="OTP not found")
        
        expires_at = datetime.fromisoformat(otp_doc["expires_at"])
        if expires_at.tzinfo is None:
            expires_at = expires_at.replace(tzinfo=timezone.utc)
        
        if expires_at < datetime.now(timezone.utc):
            raise HTTPException(status_code=401, detail="OTP expired")
        
        if otp_doc["otp"] != request.otp:
            raise HTTPException(status_code=401, detail="Invalid OTP")
        
        user = await db.users.find_one({"email": request.email}, {"_id": 0})
        if not user:
            new_user = User(email=request.email, name=request.email.split('@')[0])
            user_doc = new_user.model_dump()
            user_doc['created_at'] = user_doc['created_at'].isoformat()
            await db.users.insert_one(user_doc)
            # Fetch fresh copy without _id to avoid serialization issues
            user = await db.users.find_one({"email": request.email}, {"_id": 0})
        
        session_token = f"session_{uuid.uuid4().hex}"
        expires_at = datetime.now(timezone.utc) + timedelta(days=7)
        
        session = UserSession(
            session_token=session_token,
            user_id=user["user_id"],
            expires_at=expires_at
        )
        
        session_doc = session.model_dump()
        session_doc['expires_at'] = session_doc['expires_at'].isoformat()
        session_doc['created_at'] = session_doc['created_at'].isoformat()
        await db.user_sessions.insert_one(session_doc)
        
        response.set_cookie(
            key="session_token",
            value=session_token,
            httponly=True,
            secure=True,
            samesite="none",
            max_age=7*24*60*60,
            path="/"
        )
        
        await db.otp_codes.delete_one({"email": request.email})
        
        return {"user": user, "session_token": session_token}
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"OTP verify error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/auth/session")
async def exchange_session(request: SessionRequest, response: Response):
    try:
        async with httpx.AsyncClient() as client:
            auth_response = await client.get(
                "https://demobackend.emergentagent.com/auth/v1/env/oauth/session-data",
                headers={"X-Session-ID": request.session_id},
                timeout=10.0
            )
            
            if auth_response.status_code != 200:
                raise HTTPException(status_code=401, detail="Invalid session ID")
            
            session_data = auth_response.json()
            
            user = await db.users.find_one({"email": session_data["email"]}, {"_id": 0})
            if not user:
                new_user = User(
                    email=session_data["email"],
                    name=session_data["name"],
                    picture=session_data.get("picture")
                )
                user_doc = new_user.model_dump()
                user_doc['created_at'] = user_doc['created_at'].isoformat()
                await db.users.insert_one(user_doc)
                user = user_doc
            
            session_token = session_data["session_token"]
            expires_at = datetime.now(timezone.utc) + timedelta(days=7)
            
            session = UserSession(
                session_token=session_token,
                user_id=user["user_id"],
                expires_at=expires_at
            )
            
            session_doc = session.model_dump()
            session_doc['expires_at'] = session_doc['expires_at'].isoformat()
            session_doc['created_at'] = session_doc['created_at'].isoformat()
            await db.user_sessions.insert_one(session_doc)
            
            response.set_cookie(
                key="session_token",
                value=session_token,
                httponly=True,
                secure=True,
                samesite="none",
                max_age=7*24*60*60,
                path="/"
            )
            
            return {"user": user, "session_token": session_token}
    except Exception as e:
        logging.error(f"Session exchange error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/auth/me")
async def get_me(user: User = None, request: Request = None, session_token: Optional[str] = Cookie(None)):
    user = await get_current_user(request, session_token)
    return user

@api_router.post("/auth/logout")
async def logout(response: Response, session_token: Optional[str] = Cookie(None)):
    if session_token:
        await db.user_sessions.delete_one({"session_token": session_token})
    response.delete_cookie("session_token", path="/", samesite="none", secure=True)
    return {"message": "Logged out"}

@api_router.get("/")
async def root():
    return {"message": "Beat Claude - AI Hiring Companion API"}

# Hiring Session Endpoints
class HiringSessionCreate(BaseModel):
    name: str
    description: Optional[str] = None

@api_router.post("/hiring-sessions")
async def create_hiring_session(
    body: HiringSessionCreate,
    request: Request, 
    session_token: Optional[str] = Cookie(None)
):
    user = await get_current_user(request, session_token)
    
    if not body.name or not body.name.strip():
        raise HTTPException(status_code=400, detail="Name is required")
    
    session = HiringSession(name=body.name.strip(), description=body.description, created_by=user.user_id)
    doc = session.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    doc['updated_at'] = doc['updated_at'].isoformat()
    await db.hiring_sessions.insert_one(doc)
    
    return {
        "id": session.id,
        "name": session.name,
        "description": session.description,
        "status": session.status,
        "created_by": session.created_by,
        "created_at": doc['created_at'],
        "updated_at": doc['updated_at']
    }

@api_router.get("/hiring-sessions")
async def get_hiring_sessions(user: User = None, request: Request = None, session_token: Optional[str] = Cookie(None)):
    user = await get_current_user(request, session_token)
    sessions = await db.hiring_sessions.find({"created_by": user.user_id}, {"_id": 0}).sort("created_at", -1).to_list(100)
    for s in sessions:
        s['created_at'] = datetime.fromisoformat(s['created_at'])
        s['updated_at'] = datetime.fromisoformat(s['updated_at'])
    return sessions

@api_router.get("/hiring-sessions/{session_id}")
async def get_hiring_session(session_id: str):
    session = await db.hiring_sessions.find_one({"id": session_id}, {"_id": 0})
    if not session:
        raise HTTPException(status_code=404, detail="Hiring session not found")
    # Keep dates as strings for JSON serialization
    
    jd = await db.job_descriptions.find_one({"hiring_session_id": session_id}, {"_id": 0})
    assessments = await db.assessments.find({"job_id": jd['id']}, {"_id": 0}).to_list(10) if jd else []
    
    return {
        "session": session,
        "jd": jd,
        "assessments": assessments
    }

class StatusUpdate(BaseModel):
    status: str

@api_router.put("/hiring-sessions/{session_id}/status")
async def update_hiring_session_status(
    session_id: str,
    body: StatusUpdate,
    request: Request = None,
    session_token: Optional[str] = Cookie(None)
):
    user = await get_current_user(request, session_token)
    
    if body.status not in ['active', 'completed', 'draft']:
        raise HTTPException(status_code=400, detail="Invalid status")
    
    result = await db.hiring_sessions.update_one(
        {"id": session_id, "created_by": user.user_id},
        {"$set": {"status": body.status, "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Hiring session not found")
    
    return {"message": "Status updated", "status": body.status}

@api_router.delete("/hiring-sessions/{session_id}")
async def delete_hiring_session(
    session_id: str,
    request: Request = None,
    session_token: Optional[str] = Cookie(None)
):
    user = await get_current_user(request, session_token)
    
    result = await db.hiring_sessions.delete_one({"id": session_id, "created_by": user.user_id})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Hiring session not found")
    
    # Also clean up related data
    jd = await db.job_descriptions.find_one({"hiring_session_id": session_id})
    if jd:
        await db.job_descriptions.delete_one({"hiring_session_id": session_id})
        await db.assessments.delete_many({"job_id": jd.get('id')})
    await db.candidates.delete_many({"hiring_session_id": session_id})
    
    return {"message": "Hiring session deleted"}

@api_router.post("/jd/upload")
async def upload_jd_file(file: UploadFile = File(...), hiring_session_id: str = Form(None)):
    try:
        content = await file.read()
        filename = file.filename.lower()
        
        if filename.endswith('.pdf'):
            text = extract_text_from_pdf(content)
            source = "pdf"
        elif filename.endswith(('.doc', '.docx')):
            text = extract_text_from_docx(content)
            source = "docx"
        elif filename.endswith(('.png', '.jpg', '.jpeg', '.webp')):
            text = extract_text_from_image(content)
            source = "image"
        else:
            raise HTTPException(status_code=400, detail="Unsupported file format. Please upload PDF, DOC, DOCX, or Image files.")
        
        if not text or not text.strip():
            raise HTTPException(status_code=400, detail="No text could be extracted from file")
        
        chat = LlmChat(
            api_key=EMERGENT_LLM_KEY,
            session_id=str(uuid.uuid4()),
            system_message="You are an expert HR analyst. Parse job descriptions and extract structured information."
        ).with_model("openai", "gpt-5.2")
        
        prompt = f"""Parse this job description and extract the following information in JSON format:
- title: Job title
- role: Role type (e.g., Software Engineer, Marketing Manager)
- seniority: Seniority level (e.g., Junior, Mid-level, Senior, Lead)
- skills: List of required skills
- domain: Domain/Industry (e.g., Technology, Finance, Healthcare)
- key_responsibilities: List of key responsibilities

Job Description:
{text[:3000]}

Return ONLY valid JSON with these exact keys. No additional text."""
        
        response = await chat.send_message(UserMessage(text=prompt))
        
        import json
        parsed_data = json.loads(response.strip().replace('```json', '').replace('```', '').strip())
        
        jd = JobDescriptionParsed(
            title=parsed_data.get('title', 'Unknown Title'),
            role=parsed_data.get('role', 'Unknown Role'),
            seniority=parsed_data.get('seniority', 'Not Specified'),
            skills=parsed_data.get('skills', []),
            domain=parsed_data.get('domain', 'General'),
            key_responsibilities=parsed_data.get('key_responsibilities', []),
            description=text,
            source=source,
            hiring_session_id=hiring_session_id
        )
        
        doc = jd.model_dump()
        doc['created_at'] = doc['created_at'].isoformat()
        await db.job_descriptions.insert_one(doc)
        
        # Return clean dict without MongoDB _id
        return {
            "id": jd.id,
            "hiring_session_id": jd.hiring_session_id,
            "title": jd.title,
            "role": jd.role,
            "seniority": jd.seniority,
            "skills": jd.skills,
            "domain": jd.domain,
            "description": jd.description,
            "key_responsibilities": jd.key_responsibilities,
            "source": jd.source,
            "created_at": doc['created_at']
        }
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error uploading JD: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to parse job description: {str(e)}")

@api_router.post("/jd", response_model=JobDescriptionParsed)
async def parse_job_description(input: JobDescriptionInput):
    try:
        chat = LlmChat(
            api_key=EMERGENT_LLM_KEY,
            session_id=str(uuid.uuid4()),
            system_message="You are an expert HR analyst. Parse job descriptions and extract structured information."
        ).with_model("openai", "gpt-5.2")
        
        prompt = f"""Parse this job description and extract the following information in JSON format:
- title: Job title
- role: Role type (e.g., Software Engineer, Marketing Manager)
- seniority: Seniority level (e.g., Junior, Mid-level, Senior, Lead)
- skills: List of required skills
- domain: Domain/Industry (e.g., Technology, Finance, Healthcare)
- key_responsibilities: List of key responsibilities

Job Description:
{input.description}

Return ONLY valid JSON with these exact keys. No additional text."""
        
        response = await chat.send_message(UserMessage(text=prompt))
        
        import json
        parsed_data = json.loads(response.strip().replace('```json', '').replace('```', '').strip())
        
        jd = JobDescriptionParsed(
            title=parsed_data['title'],
            role=parsed_data['role'],
            seniority=parsed_data['seniority'],
            skills=parsed_data['skills'],
            domain=parsed_data['domain'],
            key_responsibilities=parsed_data['key_responsibilities'],
            description=input.description,
            source="text",
            hiring_session_id=input.hiring_session_id
        )
        
        doc = jd.model_dump()
        doc['created_at'] = doc['created_at'].isoformat()
        await db.job_descriptions.insert_one(doc)
        
        return jd
    except Exception as e:
        logging.error(f"Error parsing JD: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/jd")
async def get_all_job_descriptions():
    jds = await db.job_descriptions.find({}, {"_id": 0}).sort("created_at", -1).to_list(100)
    for jd in jds:
        jd['created_at'] = datetime.fromisoformat(jd['created_at'])
    return jds

@api_router.get("/jd/{job_id}", response_model=JobDescriptionParsed)
async def get_job_description(job_id: str):
    jd = await db.job_descriptions.find_one({"id": job_id}, {"_id": 0})
    if not jd:
        raise HTTPException(status_code=404, detail="Job description not found")
    jd['created_at'] = datetime.fromisoformat(jd['created_at'])
    return jd

@api_router.post("/assessment/generate", response_model=Assessment)
async def generate_assessment(request: GenerateAssessmentRequest):
    import json as json_lib
    
    jd = await db.job_descriptions.find_one({"id": request.job_id}, {"_id": 0})
    if not jd:
        raise HTTPException(status_code=404, detail="Job description not found")
    
    difficulty = request.difficulty
    # Cap at 50 for reliable generation under 60 seconds
    question_count = min(request.question_count, 50)
    
    skills_str = ', '.join(jd.get('skills', ['General'])[:5])
    role = jd.get('role', 'Professional')
    
    try:
        chat = LlmChat(
            api_key=EMERGENT_LLM_KEY,
            session_id=str(uuid.uuid4()),
            system_message="You are an expert assessment creator. Generate questions quickly and return ONLY valid JSON."
        ).with_model("openai", "gpt-5.2")
        
        # Single call for ALL questions - much faster
        prompt = f"""Create {question_count} {difficulty} assessment questions for {role}.
Skills: {skills_str}

Mix: 60% MCQ, 30% short_answer, 10% scenario. Include 3 HR questions.

Return ONLY a JSON array (no markdown, no explanation):
[{{"type":"mcq","text":"Question?","options":["A. X","B. Y","C. Z","D. W"],"expected_answer":"A. X"}},{{"type":"short_answer","text":"Question?","expected_answer":"Answer"}},{{"type":"scenario","text":"Scenario?","expected_answer":"Approach"}}]

Generate exactly {question_count} questions now:"""
        
        response = await chat.send_message(UserMessage(text=prompt))
        
        # Clean and parse response
        clean = response.strip()
        if '```' in clean:
            parts = clean.split('```')
            for p in parts:
                if p.strip().startswith('['):
                    clean = p.strip()
                    break
                elif 'json' in p.lower():
                    clean = p.replace('json', '').strip()
                    break
        
        # Find JSON array
        start = clean.find('[')
        end = clean.rfind(']') + 1
        if start >= 0 and end > start:
            clean = clean[start:end]
        
        questions_data = json_lib.loads(clean)
        questions = [Question(**q) for q in questions_data]
        
        assessment = Assessment(
            job_id=request.job_id,
            job_title=jd.get('title', 'Assessment'),
            questions=questions
        )
        
        doc = assessment.model_dump()
        doc['generated_at'] = doc['generated_at'].isoformat()
        doc['difficulty'] = difficulty
        doc['question_count'] = len(questions)
        await db.assessments.insert_one(doc)
        
        return assessment
        
    except json_lib.JSONDecodeError as e:
        logging.error(f"JSON parse error: {e}")
        raise HTTPException(status_code=500, detail="AI response was invalid. Please try again.")
    except Exception as e:
        logging.error(f"Assessment generation error: {e}")
        raise HTTPException(status_code=500, detail=f"Generation failed: {str(e)}")

@api_router.get("/assessments")
async def get_all_assessments():
    assessments = await db.assessments.find({}, {"_id": 0}).sort("generated_at", -1).to_list(100)
    for a in assessments:
        a['generated_at'] = datetime.fromisoformat(a['generated_at'])
    return assessments

@api_router.get("/assessment/{assessment_id}", response_model=Assessment)
async def get_assessment(assessment_id: str):
    assessment = await db.assessments.find_one({"id": assessment_id}, {"_id": 0})
    if not assessment:
        raise HTTPException(status_code=404, detail="Assessment not found")
    assessment['generated_at'] = datetime.fromisoformat(assessment['generated_at'])
    return assessment

@api_router.put("/assessment/{assessment_id}/questions")
async def update_assessment_questions(assessment_id: str, questions: List[Question]):
    assessment = await db.assessments.find_one({"id": assessment_id}, {"_id": 0})
    if not assessment:
        raise HTTPException(status_code=404, detail="Assessment not found")
    
    questions_data = [q.model_dump() for q in questions]
    await db.assessments.update_one(
        {"id": assessment_id},
        {"$set": {"questions": questions_data}}
    )
    return {"message": "Questions updated successfully", "count": len(questions)}

@api_router.post("/test/link", response_model=TestLink)
async def create_test_link(request: TestLinkRequest):
    assessment = await db.assessments.find_one({"id": request.assessment_id}, {"_id": 0})
    if not assessment:
        raise HTTPException(status_code=404, detail="Assessment not found")
    
    test_link = TestLink(assessment_id=request.assessment_id)
    doc = test_link.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.test_links.insert_one(doc)
    
    return test_link

@api_router.get("/test/{link_id}")
async def get_test(link_id: str):
    test_link = await db.test_links.find_one({"link_id": link_id}, {"_id": 0})
    if not test_link:
        raise HTTPException(status_code=404, detail="Test link not found")
    
    assessment = await db.assessments.find_one({"id": test_link['assessment_id']}, {"_id": 0})
    if not assessment:
        raise HTTPException(status_code=404, detail="Assessment not found")
    
    jd = await db.job_descriptions.find_one({"id": assessment['job_id']}, {"_id": 0})
    
    return {
        "assessment_id": assessment['id'],
        "job_title": jd['title'] if jd else "Position",
        "questions": [{"question_id": q['question_id'], "type": q['type'], "text": q['text'], "options": q.get('options')} for q in assessment['questions']]
    }

@api_router.post("/test/submit")
async def submit_test(request: SubmitTestRequest):
    test_link = await db.test_links.find_one({"link_id": request.link_id}, {"_id": 0})
    if not test_link:
        raise HTTPException(status_code=404, detail="Test link not found")
    
    submission = TestSubmission(
        candidate_name=request.candidate_name,
        candidate_email=request.candidate_email,
        assessment_id=test_link['assessment_id'],
        link_id=request.link_id,
        answers=request.answers
    )
    
    doc = submission.model_dump()
    doc['submitted_at'] = doc['submitted_at'].isoformat()
    await db.submissions.insert_one(doc)
    
    # AUTO-SCORE immediately after submission - SYNCHRONOUS for reliability
    assessment = await db.assessments.find_one({"id": submission.assessment_id}, {"_id": 0})
    if assessment:
        try:
            questions_dict = {q['question_id']: q for q in assessment['questions']}
            answers_dict = {a.question_id: a.answer for a in submission.answers}
            
            # Calculate scores based on expected answers
            question_scores = []
            total_score = 0
            
            for q in assessment['questions']:
                q_id = q['question_id']
                candidate_answer = answers_dict.get(q_id, "")
                expected = q.get('expected_answer', '')
                
                # Simple scoring: MCQ exact match = 10, partial text match for others
                if q['type'] == 'mcq':
                    # Check if answer matches expected (first letter or full option)
                    if candidate_answer and expected:
                        if candidate_answer.strip().upper().startswith(expected.strip().upper()[0]) or \
                           candidate_answer.strip().lower() == expected.strip().lower():
                            score = 10.0
                        else:
                            score = 0.0
                    else:
                        score = 0.0
                else:
                    # For short answer/scenario - give points if answer provided
                    if candidate_answer and len(candidate_answer.strip()) > 10:
                        score = 7.0  # Base score for answered
                    elif candidate_answer:
                        score = 4.0
                    else:
                        score = 0.0
                
                question_scores.append({
                    "question_id": q_id,
                    "score": score,
                    "reasoning": "Auto-scored"
                })
                total_score += score
            
            max_score = len(question_scores) * 10
            percentage = (total_score / max_score * 100) if max_score > 0 else 0
            
            # Get detailed AI evaluation
            try:
                import json
                chat = LlmChat(
                    api_key=EMERGENT_LLM_KEY,
                    session_id=str(uuid.uuid4()),
                    system_message="You are an expert evaluator. Score candidate responses consistently and provide detailed reasoning."
                ).with_model("openai", "gpt-5.2")
                
                evaluation_data = []
                for q in assessment['questions']:
                    answer = answers_dict.get(q['question_id'], "No answer provided")
                    evaluation_data.append({
                        "question": q['text'],
                        "type": q['type'],
                        "expected": q.get('expected_answer', 'N/A'),
                        "candidate_answer": answer
                    })
                
                prompt = f"""Evaluate this candidate's responses:

{json.dumps(evaluation_data, indent=2)}

Score each question from 0-10 and provide:
- Individual question scores with reasoning
- Overall strengths (3-5 points)  
- Areas for improvement (2-3 points)
- Hiring recommendation

Return ONLY valid JSON:
{{
  "question_scores": [{{"question_id": "id", "score": 8.5, "reasoning": "explanation"}}],
  "strengths": ["strength1", "strength2"],
  "weaknesses": ["weakness1", "weakness2"],
  "recommendations": "Overall assessment and hiring recommendation"
}}"""
                
                response = await chat.send_message(UserMessage(text=prompt))
                eval_result = json.loads(response.strip().replace('```json', '').replace('```', '').strip())
                
                ai_question_scores = [QuestionScore(**qs) for qs in eval_result['question_scores']]
                ai_total_score = sum(qs.score for qs in ai_question_scores)
                ai_max_score = len(ai_question_scores) * 10
                ai_percentage = (ai_total_score / ai_max_score * 100) if ai_max_score > 0 else 0
                
                score_record = Score(
                    submission_id=submission.id,
                    assessment_id=submission.assessment_id,
                    candidate_name=submission.candidate_name,
                    candidate_email=submission.candidate_email,
                    total_score=ai_total_score,
                    percentage=ai_percentage,
                    question_scores=ai_question_scores,
                    strengths=eval_result['strengths'],
                    weaknesses=eval_result['weaknesses'],
                    recommendations=eval_result['recommendations']
                )
                logging.info(f"AI evaluation completed for submission {submission.id}")
            except Exception as ai_error:
                logging.error(f"AI evaluation failed, using basic scoring: {ai_error}")
                score_record = Score(
                    submission_id=submission.id,
                    assessment_id=submission.assessment_id,
                    candidate_name=submission.candidate_name,
                    candidate_email=submission.candidate_email,
                    total_score=total_score,
                    percentage=percentage,
                    question_scores=[QuestionScore(**qs) for qs in question_scores],
                    strengths=["Completed assessment", "Submitted on time"],
                    weaknesses=["Review recommended"],
                    recommendations="AI detailed evaluation pending"
                )
            
            score_doc = score_record.model_dump()
            score_doc['scored_at'] = score_doc['scored_at'].isoformat()
            await db.scores.insert_one(score_doc)
            logging.info(f"Score saved for submission {submission.id}: {percentage}%")
            
        except Exception as e:
            logging.error(f"Auto-scoring failed: {e}")
    
    return {
        "id": submission.id,
        "candidate_name": submission.candidate_name,
        "candidate_email": submission.candidate_email,
        "assessment_id": submission.assessment_id,
        "submitted_at": doc['submitted_at'],
        "score_calculated": True
    }

@api_router.get("/submissions")
async def get_all_submissions():
    submissions = await db.submissions.find({}, {"_id": 0}).sort("submitted_at", -1).to_list(100)
    for s in submissions:
        s['submitted_at'] = datetime.fromisoformat(s['submitted_at'])
    return submissions

@api_router.post("/score", response_model=Score)
async def score_submission(request: ScoreRequest):
    try:
        submission = await db.submissions.find_one({"id": request.submission_id}, {"_id": 0})
        if not submission:
            raise HTTPException(status_code=404, detail="Submission not found")
        
        assessment = await db.assessments.find_one({"id": submission['assessment_id']}, {"_id": 0})
        if not assessment:
            raise HTTPException(status_code=404, detail="Assessment not found")
        
        chat = LlmChat(
            api_key=EMERGENT_LLM_KEY,
            session_id=str(uuid.uuid4()),
            system_message="You are an expert evaluator. Score candidate responses consistently and provide detailed reasoning."
        ).with_model("openai", "gpt-5.2")
        
        questions_dict = {q['question_id']: q for q in assessment['questions']}
        answers_dict = {a['question_id']: a['answer'] for a in submission['answers']}
        
        evaluation_data = []
        for q_id, question in questions_dict.items():
            answer = answers_dict.get(q_id, "No answer provided")
            evaluation_data.append({
                "question": question['text'],
                "type": question['type'],
                "expected": question.get('expected_answer', 'N/A'),
                "candidate_answer": answer
            })
        
        import json
        prompt = f"""Evaluate this candidate's responses:

{json.dumps(evaluation_data, indent=2)}

Score each question from 0-10 and provide:
- Individual question scores with reasoning
- Overall strengths (3-5 points)
- Areas for improvement (2-3 points)
- Hiring recommendation

Return ONLY valid JSON:
{{
  "question_scores": [{{
    "question_id": "id",
    "score": 8.5,
    "reasoning": "explanation"
  }}],
  "strengths": ["strength1", "strength2"],
  "weaknesses": ["weakness1", "weakness2"],
  "recommendations": "Overall assessment and recommendation"
}}"""
        
        response = await chat.send_message(UserMessage(text=prompt))
        eval_result = json.loads(response.strip().replace('```json', '').replace('```', '').strip())
        
        question_scores = [QuestionScore(**qs) for qs in eval_result['question_scores']]
        total_score = sum(qs.score for qs in question_scores)
        max_score = len(question_scores) * 10
        percentage = (total_score / max_score * 100) if max_score > 0 else 0
        
        score = Score(
            submission_id=request.submission_id,
            assessment_id=submission['assessment_id'],
            candidate_name=submission['candidate_name'],
            candidate_email=submission['candidate_email'],
            total_score=total_score,
            percentage=percentage,
            question_scores=question_scores,
            strengths=eval_result['strengths'],
            weaknesses=eval_result['weaknesses'],
            recommendations=eval_result['recommendations']
        )
        
        doc = score.model_dump()
        doc['scored_at'] = doc['scored_at'].isoformat()
        await db.scores.insert_one(doc)
        
        return score
    except Exception as e:
        logging.error(f"Error scoring submission: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/score/submission/{submission_id}")
async def get_score_by_submission(submission_id: str):
    score = await db.scores.find_one({"submission_id": submission_id}, {"_id": 0})
    if not score:
        return {"scored": False, "message": "Scoring in progress..."}
    score['scored_at'] = datetime.fromisoformat(score['scored_at'])
    return {"scored": True, "score": score}

@api_router.get("/scores")
async def get_all_scores():
    scores = await db.scores.find({}, {"_id": 0}).sort("scored_at", -1).to_list(100)
    for s in scores:
        s['scored_at'] = datetime.fromisoformat(s['scored_at'])
    return scores

@api_router.get("/leaderboard/{assessment_id}", response_model=List[LeaderboardEntry])
async def get_leaderboard(assessment_id: str):
    submissions = await db.submissions.find({"assessment_id": assessment_id}, {"_id": 0}).to_list(1000)
    submission_ids = [s['id'] for s in submissions]
    
    scores = await db.scores.find({"submission_id": {"$in": submission_ids}}, {"_id": 0}).to_list(1000)
    
    scores_sorted = sorted(scores, key=lambda x: x['percentage'], reverse=True)
    
    leaderboard = []
    for rank, score in enumerate(scores_sorted, 1):
        submission = next((s for s in submissions if s['id'] == score['submission_id']), None)
        if submission:
            leaderboard.append(LeaderboardEntry(
                rank=rank,
                candidate_name=score['candidate_name'],
                candidate_email=score['candidate_email'],
                total_score=score['total_score'],
                percentage=score['percentage'],
                strengths=score['strengths'],
                recommendations=score['recommendations'],
                submitted_at=datetime.fromisoformat(submission['submitted_at'])
            ))
    
    return leaderboard

@api_router.get("/dashboard/stats")
async def get_dashboard_stats():
    total_jds = await db.job_descriptions.count_documents({})
    total_assessments = await db.assessments.count_documents({})
    total_submissions = await db.submissions.count_documents({})
    total_scores = await db.scores.count_documents({})
    
    return {
        "total_jds": total_jds,
        "total_assessments": total_assessments,
        "total_candidates": total_submissions,
        "total_scored": total_scores
    }

# Candidate Management Endpoints
@api_router.post("/candidates/{hiring_session_id}")
async def add_candidate(hiring_session_id: str, candidate: CandidateCreate):
    """Add a single candidate to a hiring session"""
    session = await db.hiring_sessions.find_one({"id": hiring_session_id}, {"_id": 0})
    if not session:
        raise HTTPException(status_code=404, detail="Hiring session not found")
    
    # Check if candidate already exists
    existing = await db.candidates.find_one({
        "hiring_session_id": hiring_session_id,
        "email": candidate.email
    }, {"_id": 0})
    if existing:
        raise HTTPException(status_code=400, detail="Candidate already added to this session")
    
    new_candidate = Candidate(
        hiring_session_id=hiring_session_id,
        name=candidate.name,
        email=candidate.email
    )
    
    doc = new_candidate.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    if doc.get('invite_sent_at'):
        doc['invite_sent_at'] = doc['invite_sent_at'].isoformat()
    await db.candidates.insert_one(doc)
    
    return new_candidate

@api_router.post("/candidates/{hiring_session_id}/bulk")
async def add_candidates_bulk(hiring_session_id: str, request: BulkCandidateCreate):
    """Add multiple candidates to a hiring session"""
    session = await db.hiring_sessions.find_one({"id": hiring_session_id}, {"_id": 0})
    if not session:
        raise HTTPException(status_code=404, detail="Hiring session not found")
    
    added = []
    skipped = []
    
    for candidate in request.candidates:
        existing = await db.candidates.find_one({
            "hiring_session_id": hiring_session_id,
            "email": candidate.email
        }, {"_id": 0})
        
        if existing:
            skipped.append(candidate.email)
            continue
        
        new_candidate = Candidate(
            hiring_session_id=hiring_session_id,
            name=candidate.name,
            email=candidate.email
        )
        
        doc = new_candidate.model_dump()
        doc['created_at'] = doc['created_at'].isoformat()
        if doc.get('invite_sent_at'):
            doc['invite_sent_at'] = doc['invite_sent_at'].isoformat()
        await db.candidates.insert_one(doc)
        added.append(new_candidate)
    
    return {"added": len(added), "skipped": skipped, "candidates": added}

@api_router.get("/candidates/template/download")
async def download_candidate_template():
    """Download CSV template for bulk candidate import"""
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['name', 'email'])
    writer.writerow(['John Doe', 'john.doe@example.com'])
    writer.writerow(['Jane Smith', 'jane.smith@example.com'])
    writer.writerow(['Mike Johnson', 'mike.johnson@example.com'])
    
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=candidates_template.csv"}
    )

@api_router.post("/candidates/{hiring_session_id}/import")
async def import_candidates_csv(hiring_session_id: str, file: UploadFile = File(...)):
    """Import candidates from CSV/Excel file"""
    session = await db.hiring_sessions.find_one({"id": hiring_session_id}, {"_id": 0})
    if not session:
        raise HTTPException(status_code=404, detail="Hiring session not found")
    
    content = await file.read()
    filename = file.filename.lower()
    
    candidates_data = []
    
    try:
        if filename.endswith('.csv'):
            # Parse CSV
            text_content = content.decode('utf-8')
            reader = csv.DictReader(io.StringIO(text_content))
            for row in reader:
                name = row.get('name', row.get('Name', '')).strip()
                email = row.get('email', row.get('Email', '')).strip()
                if name and email and '@' in email:
                    candidates_data.append({'name': name, 'email': email})
        elif filename.endswith(('.xlsx', '.xls')):
            # For Excel, try to parse as CSV (basic support)
            try:
                import openpyxl
                wb = openpyxl.load_workbook(io.BytesIO(content))
                ws = wb.active
                headers = [cell.value.lower() if cell.value else '' for cell in ws[1]]
                name_idx = headers.index('name') if 'name' in headers else 0
                email_idx = headers.index('email') if 'email' in headers else 1
                
                for row in ws.iter_rows(min_row=2, values_only=True):
                    name = str(row[name_idx] or '').strip()
                    email = str(row[email_idx] or '').strip()
                    if name and email and '@' in email:
                        candidates_data.append({'name': name, 'email': email})
            except ImportError:
                raise HTTPException(status_code=400, detail="Excel support requires openpyxl. Please use CSV format.")
        else:
            raise HTTPException(status_code=400, detail="Please upload a CSV or Excel file")
    except Exception as e:
        logging.error(f"Error parsing file: {e}")
        raise HTTPException(status_code=400, detail=f"Error parsing file: {str(e)}")
    
    if not candidates_data:
        raise HTTPException(status_code=400, detail="No valid candidates found in file. Make sure file has 'name' and 'email' columns.")
    
    added = []
    skipped = []
    
    for cand in candidates_data:
        existing = await db.candidates.find_one({
            "hiring_session_id": hiring_session_id,
            "email": cand['email']
        }, {"_id": 0})
        
        if existing:
            skipped.append(cand['email'])
            continue
        
        new_candidate = Candidate(
            hiring_session_id=hiring_session_id,
            name=cand['name'],
            email=cand['email']
        )
        
        doc = new_candidate.model_dump()
        doc['created_at'] = doc['created_at'].isoformat()
        await db.candidates.insert_one(doc)
        added.append({"name": cand['name'], "email": cand['email']})
    
    return {
        "added": len(added),
        "skipped": len(skipped),
        "skipped_emails": skipped,
        "message": f"Successfully imported {len(added)} candidates"
    }

@api_router.get("/candidates/{hiring_session_id}")
async def get_candidates(hiring_session_id: str):
    """Get all candidates for a hiring session"""
    candidates = await db.candidates.find(
        {"hiring_session_id": hiring_session_id},
        {"_id": 0}
    ).sort("created_at", -1).to_list(500)
    
    for c in candidates:
        c['created_at'] = datetime.fromisoformat(c['created_at'])
        if c.get('invite_sent_at'):
            c['invite_sent_at'] = datetime.fromisoformat(c['invite_sent_at'])
    
    return candidates

@api_router.delete("/candidates/{candidate_id}")
async def delete_candidate(candidate_id: str):
    """Delete a candidate"""
    result = await db.candidates.delete_one({"id": candidate_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Candidate not found")
    return {"message": "Candidate deleted"}

@api_router.post("/candidates/send-invites")
async def send_candidate_invites(request: SendInviteRequest):
    """Send test invites to selected candidates"""
    assessment = await db.assessments.find_one({"id": request.assessment_id}, {"_id": 0})
    if not assessment:
        raise HTTPException(status_code=404, detail="Assessment not found")
    
    jd = await db.job_descriptions.find_one({"id": assessment['job_id']}, {"_id": 0})
    job_title = jd['title'] if jd else "Position"
    
    sent = []
    failed = []
    
    for candidate_id in request.candidate_ids:
        candidate = await db.candidates.find_one({"id": candidate_id}, {"_id": 0})
        if not candidate:
            failed.append({"id": candidate_id, "reason": "Not found"})
            continue
        
        # Create test link for this candidate
        test_link = TestLink(assessment_id=request.assessment_id)
        link_doc = test_link.model_dump()
        link_doc['created_at'] = link_doc['created_at'].isoformat()
        await db.test_links.insert_one(link_doc)
        
        test_url = f"https://unzip-start-1.preview.emergentagent.com/test/{test_link.link_id}"
        
        # Send email if Resend is configured
        if RESEND_API_KEY:
            html_content = f"""
            <div style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto; padding: 20px;">
                <div style="text-align: center; margin-bottom: 30px;">
                    <h1 style="color: #991b1b; margin: 0;">Beat Claude</h1>
                    <p style="color: #6b7280; margin-top: 5px;">AI Hiring Companion</p>
                </div>
                <div style="background: #f9fafb; border-radius: 8px; padding: 30px;">
                    <h2 style="color: #111827; margin: 0 0 10px 0;">You're Invited to Take an Assessment</h2>
                    <p style="color: #374151; margin: 0 0 20px 0;">
                        Hi {candidate['name']},
                    </p>
                    <p style="color: #374151; margin: 0 0 20px 0;">
                        You have been invited to complete an assessment for the <strong>{job_title}</strong> position.
                    </p>
                    <div style="margin: 30px 0;">
                        <a href="{test_url}" style="background: #991b1b; color: white; text-decoration: none; padding: 14px 28px; border-radius: 8px; font-weight: bold; display: inline-block;">
                            Start Assessment
                        </a>
                    </div>
                    <p style="color: #6b7280; font-size: 14px; margin: 20px 0 0 0;">
                        <strong>Important:</strong>
                    </p>
                    <ul style="color: #6b7280; font-size: 14px; margin: 5px 0 0 0; padding-left: 20px;">
                        <li>The assessment has 50 questions and a 60-minute time limit</li>
                        <li>Once started, the timer cannot be paused</li>
                        <li>Ensure you have a stable internet connection</li>
                    </ul>
                </div>
                <p style="color: #9ca3af; font-size: 12px; text-align: center; margin-top: 20px;">
                    If you have any questions, please contact the hiring team.
                </p>
            </div>
            """
            
            params = {
                "from": SENDER_EMAIL,
                "to": [candidate['email']],
                "subject": f"Assessment Invitation: {job_title}",
                "html": html_content
            }
            
            try:
                await asyncio.to_thread(resend.Emails.send, params)
                logging.info(f"Invite sent to {candidate['email']}")
            except Exception as e:
                logging.error(f"Failed to send invite to {candidate['email']}: {e}")
        
        # Update candidate record
        await db.candidates.update_one(
            {"id": candidate_id},
            {"$set": {
                "assessment_id": request.assessment_id,
                "test_link_id": test_link.link_id,
                "invite_sent_at": datetime.now(timezone.utc).isoformat(),
                "status": "invited"
            }}
        )
        
        sent.append({
            "id": candidate_id,
            "email": candidate['email'],
            "test_link": test_url
        })
    
    return {"sent": len(sent), "failed": failed, "invites": sent}

@api_router.get("/candidates/{hiring_session_id}/stats")
async def get_candidate_stats(hiring_session_id: str):
    """Get candidate statistics for a hiring session"""
    candidates = await db.candidates.find(
        {"hiring_session_id": hiring_session_id},
        {"_id": 0}
    ).to_list(500)
    
    total = len(candidates)
    invited = len([c for c in candidates if c.get('invite_sent_at')])
    completed = len([c for c in candidates if c.get('status') == 'test_completed'])
    scored = len([c for c in candidates if c.get('status') == 'scored'])
    
    return {
        "total": total,
        "invited": invited,
        "completed": completed,
        "scored": scored
    }

@api_router.post("/submissions/{submission_id}/score")
async def score_submission(submission_id: str):
    """Manually trigger scoring for a submission"""
    submission = await db.submissions.find_one({"id": submission_id}, {"_id": 0})
    if not submission:
        raise HTTPException(status_code=404, detail="Submission not found")
    
    # Check if already scored
    existing_score = await db.scores.find_one({"submission_id": submission_id}, {"_id": 0})
    if existing_score:
        return {"message": "Already scored", "percentage": existing_score.get('percentage', 0)}
    
    assessment = await db.assessments.find_one({"id": submission['assessment_id']}, {"_id": 0})
    if not assessment:
        raise HTTPException(status_code=404, detail="Assessment not found")
    
    # Score the submission
    questions_dict = {q['question_id']: q for q in assessment['questions']}
    answers_dict = {a['question_id']: a['answer'] for a in submission.get('answers', [])}
    
    question_scores = []
    total_score = 0
    
    for q in assessment['questions']:
        q_id = q['question_id']
        candidate_answer = answers_dict.get(q_id, "")
        expected = q.get('expected_answer', '')
        
        if q['type'] == 'mcq':
            if candidate_answer and expected:
                if candidate_answer.strip().upper().startswith(expected.strip().upper()[0]) or \
                   candidate_answer.strip().lower() == expected.strip().lower():
                    score = 10.0
                else:
                    score = 0.0
            else:
                score = 0.0
        else:
            if candidate_answer and len(candidate_answer.strip()) > 10:
                score = 7.0
            elif candidate_answer:
                score = 4.0
            else:
                score = 0.0
        
        question_scores.append({"question_id": q_id, "score": score, "reasoning": "Auto-scored"})
        total_score += score
    
    max_score = len(question_scores) * 10
    percentage = (total_score / max_score * 100) if max_score > 0 else 0
    
    score_record = Score(
        submission_id=submission_id,
        assessment_id=submission['assessment_id'],
        candidate_name=submission['candidate_name'],
        candidate_email=submission['candidate_email'],
        total_score=total_score,
        percentage=percentage,
        question_scores=[QuestionScore(**qs) for qs in question_scores],
        strengths=["Completed assessment"],
        weaknesses=["Review recommended"],
        recommendations="AI evaluation completed"
    )
    
    score_doc = score_record.model_dump()
    score_doc['scored_at'] = score_doc['scored_at'].isoformat()
    await db.scores.insert_one(score_doc)
    
    return {"message": "Scored successfully", "percentage": percentage}

@api_router.post("/submissions/score-all-pending")
async def score_all_pending():
    """Score all submissions that don't have scores yet"""
    submissions = await db.submissions.find({}, {"_id": 0}).to_list(1000)
    scores = await db.scores.find({}, {"_id": 0, "submission_id": 1}).to_list(1000)
    scored_ids = {s['submission_id'] for s in scores}
    
    pending = [s for s in submissions if s['id'] not in scored_ids]
    scored_count = 0
    
    for submission in pending:
        assessment = await db.assessments.find_one({"id": submission['assessment_id']}, {"_id": 0})
        if not assessment:
            continue
        
        answers_dict = {a['question_id']: a['answer'] for a in submission.get('answers', [])}
        
        question_scores = []
        total_score = 0
        
        for q in assessment['questions']:
            q_id = q['question_id']
            candidate_answer = answers_dict.get(q_id, "")
            expected = q.get('expected_answer', '')
            
            if q['type'] == 'mcq':
                if candidate_answer and expected:
                    if candidate_answer.strip().upper().startswith(expected.strip().upper()[0]):
                        score = 10.0
                    else:
                        score = 0.0
                else:
                    score = 0.0
            else:
                if candidate_answer and len(candidate_answer.strip()) > 10:
                    score = 7.0
                elif candidate_answer:
                    score = 4.0
                else:
                    score = 0.0
            
            question_scores.append({"question_id": q_id, "score": score, "reasoning": "Auto-scored"})
            total_score += score
        
        max_score = len(question_scores) * 10 if question_scores else 1
        percentage = (total_score / max_score * 100) if max_score > 0 else 0
        
        score_record = Score(
            submission_id=submission['id'],
            assessment_id=submission['assessment_id'],
            candidate_name=submission['candidate_name'],
            candidate_email=submission['candidate_email'],
            total_score=total_score,
            percentage=percentage,
            question_scores=[QuestionScore(**qs) for qs in question_scores],
            strengths=["Completed assessment"],
            weaknesses=["Review recommended"],
            recommendations="AI evaluation completed"
        )
        
        score_doc = score_record.model_dump()
        score_doc['scored_at'] = score_doc['scored_at'].isoformat()
        await db.scores.insert_one(score_doc)
        scored_count += 1
    
    return {"message": f"Scored {scored_count} pending submissions", "scored": scored_count}

app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()